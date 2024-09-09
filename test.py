import openpyxl
import json
import re

# Excelファイルの読み込み
load_book = openpyxl.load_workbook('/Users/takuyaishikawa/Desktop/OPE室/asasa.xlsm')
sheet = load_book['最新リスト']

# データを格納するリスト
data_list = []

# 正規表現パターン（数字のみを抽出する）
pattern = re.compile(r'\d+')

# ExcelシートのA列（barcode）とB列（name）のデータを抽出
# 1行目も含めて処理するが、A列が文字列のみの場合は無視する
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
    barcode, name = row
    
    # barcodeがNoneでない場合のみ処理
    if barcode and name:
        # barcodeを文字列として扱う
        barcode_str = str(barcode)
        
        # barcodeから数字のみを抽出
        clean_barcode = ''.join(pattern.findall(barcode_str))
        
        # 抽出された数字が存在しない場合はスキップ
        if not clean_barcode:
            continue  # 数字が含まれていないのでこの行を無視

        # 抽出された数字部分のみをリストに追加
        data_list.append({
            "barcode": clean_barcode,  # 数字部分のみ
            "name": name
        })

# 抽出したデータをJSON形式に変換
json_data = {
    "data": data_list
}

# JSONファイルに保存
json_path = '/Users/takuyaishikawa/Desktop/OPE室/output.json'
with open(json_path, 'w', encoding='utf-8') as json_file:
    json.dump(json_data, json_file, ensure_ascii=False, indent=4)

print(f"JSONファイルが出力されました: {json_path}")
