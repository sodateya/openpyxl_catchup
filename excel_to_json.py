import openpyxl
import json
import re
import os
import tkinter as tk
from tkinter import filedialog

# ファイル選択ダイアログを表示してExcelファイルを選択させる
root = tk.Tk()
root.withdraw()  # Tkinterウィンドウを非表示にする

# ファイル選択ダイアログを表示
excel_file_path = filedialog.askopenfilename(
    title="Excelファイルを選択してください",
    filetypes=[("Excel files", "*.xlsx *.xlsm")]
)

# ファイルが選択されなかった場合の処理
if not excel_file_path:
    print("ファイルが選択されませんでした。")
    exit()

# Excelファイルのディレクトリとファイル名を取得
file_dir = os.path.dirname(excel_file_path)
file_name = os.path.splitext(os.path.basename(excel_file_path))[0]

# 正規表現パターン（数字のみを抽出する）
pattern = re.compile(r'\d+')

# Excelファイルの読み込み
try:
    load_book = openpyxl.load_workbook(excel_file_path)
except Exception as e:
    print(f"Excelファイルの読み込みに失敗しました: {e}")
    exit()

# シート「最新リスト」の存在確認
if "最新リスト" not in load_book.sheetnames:
    print("シート名『最新リスト』が見つかりませんでした。")
    exit()

sheet = load_book['最新リスト']

# データを格納するリスト
data_list = []

# ExcelシートのA列（barcode）とB列（name）のデータを抽出
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
    barcode, name = row
    
    # barcodeがNoneでない場合のみ処理
    if barcode and name:
        # barcodeを文字列として扱う
        barcode_str = str(barcode)
        
        # barcodeから数字のみを抽出
        clean_barcode = ''.join(pattern.findall(barcode_str))
        
        # 抽出された数字が存在しない場合はスキップ
        if not clean_barcode:
            continue

        # 抽出された数字部分のみをリストに追加
        data_list.append({
            "barcode": clean_barcode,  # 数字部分のみ
            "name": name
        })

# 抽出したデータをJSON形式に変換
json_data = {
    "data": data_list
}

# JSONファイルに保存
json_path = os.path.join(file_dir, f"{file_name}_output.json")
with open(json_path, 'w', encoding='utf-8') as json_file:
    json.dump(json_data, json_file, ensure_ascii=False, indent=4)

print(f"JSONファイルが出力されました: {json_path}")
